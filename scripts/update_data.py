"""
update_data.py

Reads MATA's raw exported CSV/XLSX reports out of the /data folder, computes
the same KPIs shown on the executive dashboard, and writes the result to
data.json at the repo root.

This is the piece that would eventually be replaced by a real API call
(e.g. Trapeze, Genfare, Navineo) once you have direct system access -- for
now it re-parses whatever exports get dropped into /data, which is a
completely valid and common "integration" pattern when a system doesn't
expose a live API yet.

Expected input files (place these in the /data folder, replacing them each
time you have a fresh export):
  - data/mayors_dashboard.csv          (Mayor's Dashboard FY export)
  - data/route_rankings.csv            (Fixed Route Ridership Rankings export)
  - data/stats_update.xlsx             (MATA Stats Update workbook)

Requires: openpyxl  (pip install openpyxl --break-system-packages)

Run manually with:
    python scripts/update_data.py

Language: Python 3
"""

# --- Standard library imports ---
import csv                              # lets us read CSV files row-by-row instead of
                                         # writing our own comma-splitting logic
import json                             # converts our Python dictionaries/lists into
                                         # the JSON text format the dashboard reads
from pathlib import Path                # a modern, object-oriented way to work with
                                         # file paths (instead of raw strings)
from datetime import datetime, timezone # used to timestamp when this script last ran

# --- Third-party import ---
import openpyxl   # the library that knows how to open .xlsx Excel files;
                  # Python's standard library can't read Excel format on its own

# ---------------------------------------------------------------------------
# PATH SETUP
# ---------------------------------------------------------------------------

# __file__ is this script's own path. .resolve() makes it an absolute path,
# and .parent.parent walks up two folder levels (out of /scripts, to the repo
# root) so this script works no matter where it's run from.
REPO_ROOT = Path(__file__).resolve().parent.parent

DATA_DIR = REPO_ROOT / "data"          # the "/" here isn't division -- Path objects
                                        # overload it to join folder names together
OUTPUT_FILE = REPO_ROOT / "data.json"  # where the final result gets written

# The fiscal year runs July -> June, so this is the month order used
# everywhere we display "monthly" data on the dashboard.
MONTHS = ["Jul", "Aug", "Sep", "Oct", "Nov", "Dec",
          "Jan", "Feb", "Mar", "Apr", "May", "Jun"]


# ---------------------------------------------------------------------------
# SMALL HELPER FUNCTIONS
# ---------------------------------------------------------------------------

def clean_number(raw):
    """
    Turn strings like ' 2,549,260 ' or '61.2%' into a plain float.
    Returns None if the cell was blank (some months haven't reported yet).

    Why this is needed: numbers straight out of an Excel/CSV export often
    come as text with commas, percent signs, or stray whitespace baked in --
    "2,549,260" can't be turned into a number until those are stripped out.
    """
    if raw is None:
        return None

    # Remove thousands-separator commas, stray quote characters, and
    # leading/trailing whitespace (spreadsheet exports love adding these).
    text = raw.strip().replace(",", "").replace('"', "")

    if text == "":
        # An empty cell means "no data reported for this month" -- keep
        # that distinction (None) rather than treating it as zero.
        return None

    is_percent = text.endswith("%")   # noted but not currently used further;
                                       # kept in case future logic needs to
                                       # distinguish percentages from counts
    text = text.rstrip("%")           # drop the % sign so float() can parse it

    try:
        value = float(text)
    except ValueError:
        # If it still isn't a valid number after cleaning (e.g. stray text),
        # treat it the same as "no data" rather than crashing the script.
        return None

    return value


def load_csv_rows(path):
    """
    Opens a CSV file and returns its contents as a list of rows, where each
    row is itself a list of cell values (strings).

    encoding="cp1252" matters here: these files were exported from Windows
    Excel, which uses this character encoding by default. Using the wrong
    encoding (like plain "utf-8") would corrupt special characters such as
    the en-dash ("–") used in route names like "50 – Poplar".
    """
    with open(path, encoding="cp1252", newline="") as f:
        return list(csv.reader(f))


def find_row(rows, label_contains):
    """
    Scans through every row and returns the first one whose first cell
    contains the given text (case-insensitive).

    This is how we locate a specific line in a messy report -- e.g. finding
    the row for "Ridership - MotorBus" without needing to know its exact
    row number, which could shift if MATA reorders the spreadsheet later.
    """
    for row in rows:
        if row and label_contains.lower() in row[0].lower():
            return row
    return None


# ---------------------------------------------------------------------------
# PARSER 1: Mayor's Dashboard CSV (monthly KPIs -- ridership + on-time %)
# ---------------------------------------------------------------------------

def parse_mayors_dashboard(path):
    rows = load_csv_rows(path)

    def monthly_and_total(label):
        """
        Small helper *nested inside* parse_mayors_dashboard, since it's only
        ever used here. Given a row label (like "Ridership - MotorBus"),
        finds that row and pulls out:
          - the 12 monthly values (columns 1 through 12, since column 0 is
            the label itself)
          - the yearly total (column 14, based on this report's layout)
        """
        row = find_row(rows, label)
        if row is None:
            return None, None
        monthly = [clean_number(v) for v in row[1:13]]   # columns 1-12 = Jul-Jun
        total = clean_number(row[14]) if len(row) > 14 else None
        return monthly, total

    # Pull each ridership category by mode of transportation.
    motorbus_ridership, motorbus_total = monthly_and_total("Ridership - MotorBus")
    trolley_ridership, trolley_total = monthly_and_total("Ridership - Steel Rail")
    paratransit_ridership, paratransit_total = monthly_and_total("Ridership - Demand Response")
    ready1_monthly, ready1_total = monthly_and_total("Ready! Zone 1")
    ready2_monthly, ready2_total = monthly_and_total("Ready! Zone 2")
    ready3_monthly, ready3_total = monthly_and_total("Ready! Zone 3")
    groove_monthly, groove_total = monthly_and_total("Groove")

    # On-time performance (OTP) rows work a little differently from
    # ridership rows -- they have a "Goal" value in column 13 instead of
    # a running total, since a percentage can't be summed across months.
    otp_fixed_row = find_row(rows, "On Time Performance- MotorBus")
    otp_para_row = find_row(rows, "On Time Performance - Demand Response")

    otp_fixed_monthly = [clean_number(v) for v in otp_fixed_row[1:13]] if otp_fixed_row else []
    otp_fixed_goal = clean_number(otp_fixed_row[13]) if otp_fixed_row else None
    otp_para_monthly = [clean_number(v) for v in otp_para_row[1:13]] if otp_para_row else []
    otp_para_goal = clean_number(otp_para_row[13]) if otp_para_row else None

    def latest(monthly):
        """
        Returns the most recent non-blank value in a list of monthly
        numbers. Since the fiscal year is still in progress, later months
        may still be blank (not reported yet) -- this finds the last month
        that actually HAS a number, rather than assuming month 12 is filled.
        """
        vals = [v for v in monthly if v is not None]
        return vals[-1] if vals else None

    # Add up every mode's yearly total to get one combined ridership number.
    # Filtering out None first means a mode with missing data doesn't break
    # the whole sum.
    totals = [t for t in [motorbus_total, trolley_total, paratransit_total,
                          ready1_total, ready2_total, ready3_total, groove_total]
              if t is not None]

    # Return everything as one dictionary -- this becomes part of the final
    # JSON structure the dashboard reads.
    return {
        "totalRidershipYTD": sum(totals) if totals else None,
        "otpFixedRoute": latest(otp_fixed_monthly),
        "otpFixedRouteGoal": otp_fixed_goal,
        "otpParatransit": latest(otp_para_monthly),
        "otpParatransitGoal": otp_para_goal,
        "ridershipByMode": {
            "months": MONTHS,
            "motorbus": motorbus_ridership or [],     # "or []" avoids sending
            "trolley": trolley_ridership or [],        # None to the dashboard
            "paratransit": paratransit_ridership or [],# if a row wasn't found
        },
        "otpTrend": {
            "months": MONTHS,
            "fixedRoute": otp_fixed_monthly,
            "paratransit": otp_para_monthly,
        },
    }


# ---------------------------------------------------------------------------
# PARSER 2: Fixed Route Ridership Rankings CSV (top routes by ridership)
# ---------------------------------------------------------------------------

def parse_route_rankings(path, fiscal_year_label="FY26"):
    """
    This report stacks multiple fiscal years in one file, one block after
    another (FY22, FY23, FY24... FY26), each with its own little header row.
    This function finds the block for the fiscal year we care about, reads
    just that block, and returns the top 10 routes by total ridership.
    """
    rows = load_csv_rows(path)

    # Step 1: find which row starts the FY26 section by searching for a
    # row whose first cell mentions "FY26" (e.g. a header like "FY26 Data").
    start_index = None
    for i, row in enumerate(rows):
        if row and fiscal_year_label.lower() in row[0].lower():
            start_index = i + 1   # the actual route data starts right after
                                   # that header row
            break

    if start_index is None:
        # If we can't find an FY26 section at all, return an empty list
        # rather than crashing -- lets the rest of the dashboard still work.
        return []

    routes = []
    for row in rows[start_index:]:
        if not row or not row[0].strip():
            # A blank first cell signals we've reached the end of this
            # fiscal year's block (the next block, if any, starts further down).
            break
        name = row[0].strip()
        total = clean_number(row[13]) if len(row) > 13 else None
        if total is not None:
            routes.append({"route": name, "riders": int(total)})

    # Sort so the highest-ridership route comes first, then keep only the
    # top 10 -- that's all the dashboard's table needs to display.
    routes.sort(key=lambda r: r["riders"], reverse=True)
    return routes[:10]


# ---------------------------------------------------------------------------
# PARSER 3: MATA Stats Update workbook (route-level on-time performance)
# ---------------------------------------------------------------------------

def parse_stats_update(path):
    """
    Reads the 'OTP by Route' and '3mon - OTP' sheets and returns per-route
    on-time performance for the most recent reporting month, plus a trend
    arrow (up/down) showing direction over the last 3 months.
    """
    # data_only=True is important: without it, openpyxl would return the
    # *formulas* stored in each cell (like "=AVERAGE(B2:B4)") instead of
    # the calculated numbers Excel actually displays.
    wb = openpyxl.load_workbook(path, data_only=True)
    otp_sheet = wb["OTP by Route"]

    header_row = 3      # row 3 contains the month/date headers for each column
    data_start_row = 4  # route data begins on row 4

    # Grab every cell in the header row so we can look up "which month is
    # column 7?" later on.
    dates = [cell.value for cell in otp_sheet[header_row]]

    latest_by_route = {}
    for r in range(data_start_row, otp_sheet.max_row + 1):
        route = otp_sheet.cell(row=r, column=1).value
        if not route or str(route).strip().lower() == "total":
            continue   # skip blank rows and any "Total" summary row
        route = str(route).strip()

        # Scan the row from right to left (most recent month first). The
        # first non-empty cell we hit is that route's latest reported OTP --
        # this handles routes that stopped reporting early (discontinued
        # routes) as well as the current month simply not being filled in yet.
        for c in range(otp_sheet.max_column, 1, -1):
            value = otp_sheet.cell(row=r, column=c).value
            if value is not None:
                latest_by_route[route] = {
                    "otp": round(value * 100, 1),   # stored as a decimal
                                                     # fraction (0.612), so
                                                     # convert to a percent
                    "asOf": dates[c - 1].strftime("%Y-%m-%d") if dates[c - 1] else None,
                }
                break   # stop scanning this row -- we found what we needed

    if not latest_by_route:
        return []

    # Some routes (like discontinued ones) stopped reporting months or
    # years ago. We only want routes that are CURRENTLY active, so we find
    # the most recent date seen across all routes, then keep only the
    # routes whose latest entry matches that date.
    most_recent = max(v["asOf"] for v in latest_by_route.values() if v["asOf"])
    active_routes = {r: v for r, v in latest_by_route.items() if v["asOf"] == most_recent}

    # Now merge in the 3-month trend arrow (↑ or ↓) from the second sheet,
    # matching routes by name between the two sheets.
    if "3mon - OTP" in wb.sheetnames:
        trend_sheet = wb["3mon - OTP"]
        for r in range(2, trend_sheet.max_row + 1):
            route = trend_sheet.cell(row=r, column=1).value
            if not route:
                continue
            route = str(route).strip()
            if route in active_routes:
                arrow = trend_sheet.cell(row=r, column=6).value
                active_routes[route]["trend"] = arrow

    # Convert from a dictionary keyed by route name into a plain list of
    # {route, otp, trend} objects -- easier for the dashboard's JavaScript
    # to loop over directly.
    result = [
        {"route": route, "otp": v["otp"], "trend": v.get("trend", "")}
        for route, v in active_routes.items()
    ]

    # Sort worst-to-best so underperforming routes are the first thing a
    # reviewer sees -- the whole point of an executive dashboard is
    # surfacing what needs attention.
    result.sort(key=lambda r: r["otp"])
    return result


# ---------------------------------------------------------------------------
# MAIN ENTRY POINT
# ---------------------------------------------------------------------------

def main():
    mayors_csv = DATA_DIR / "mayors_dashboard.csv"
    rankings_csv = DATA_DIR / "route_rankings.csv"
    stats_xlsx = DATA_DIR / "stats_update.xlsx"

    # The two CSVs are considered required -- if they're missing, stop
    # immediately with a clear error rather than silently producing a
    # broken/empty dashboard.
    if not mayors_csv.exists() or not rankings_csv.exists():
        raise FileNotFoundError(
            "Expected data/mayors_dashboard.csv and data/route_rankings.csv "
            "to exist. Drop your latest CSV exports into the /data folder."
        )

    kpi_data = parse_mayors_dashboard(mayors_csv)
    top_routes = parse_route_rankings(rankings_csv)

    # The xlsx file is treated as optional -- if it's not there yet, the
    # dashboard still works, just without the route-level OTP table.
    otp_by_route = parse_stats_update(stats_xlsx) if stats_xlsx.exists() else []

    # {**kpi_data, ...} "unpacks" all the key/value pairs from kpi_data
    # directly into this new dictionary, so the final JSON is one flat
    # object instead of being nested inside a "kpiData" sub-key.
    output = {
        "lastUpdated": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"),
        **kpi_data,
        "topRoutes": top_routes,
        "otpByRoute": otp_by_route,
    }

    # indent=2 makes the JSON file human-readable (nicely spaced), which
    # helps when you're checking it in GitHub or debugging by eye.
    OUTPUT_FILE.write_text(json.dumps(output, indent=2))

    print(f"Wrote {OUTPUT_FILE} with {len(top_routes)} top routes and "
          f"{len(otp_by_route)} route-level OTP entries.")


# This check means "only run main() if this file is executed directly"
# (e.g. `python scripts/update_data.py`) -- not if some other script
# imports functions from this file without wanting it to auto-run.
if __name__ == "__main__":
    main()
